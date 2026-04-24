"""Writes OneStop quantities into a copy of the active GM template.

openpyxl's default load → modify → save path preserves cell styles, merged
ranges, column widths, images, and print settings. We only touch the ORDER
cells of matched rows, so the output is byte-equivalent to the input
everywhere else (as close to byte-equivalent as openpyxl can manage — there
are known quirks around VBA-enabled files, but the GM template is a plain
.xlsx, not .xlsm).

Caller responsibilities:
  - Decide the target item# / sheet / side / row for each write.
  - This module does NOT do matching. It just writes.
"""
from __future__ import annotations

import shutil
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook


# ORDER column is column C on the left block and column H on the right block.
ORDER_COL = {"left": "C", "right": "H"}


@dataclass(frozen=True)
class OrderWrite:
    sheet: str
    side: str          # "left" or "right"
    row_index: int     # 1-based row on the sheet
    item_no: int       # for sanity check before we write
    qty: int


def write_quantities(
    template_path: Path | str,
    output_path: Path | str,
    writes: list[OrderWrite],
) -> int:
    """Copy the template to ``output_path`` and set ORDER cells in place.

    Returns the number of cells written. Raises if a sanity check fails —
    we'd rather hard-error than write a quantity to the wrong row.
    """
    template_path = Path(template_path)
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    # Start from a fresh filesystem copy so load-modify-save operates on the
    # output file and the template stays pristine.
    shutil.copyfile(template_path, output_path)

    wb = load_workbook(output_path)  # NOT read_only — we're modifying
    written = 0
    for w in writes:
        if w.sheet not in wb.sheetnames:
            raise ValueError(f"GM template has no sheet {w.sheet!r}")
        if w.side not in ORDER_COL:
            raise ValueError(f"invalid side {w.side!r} (expected 'left'/'right')")
        ws = wb[w.sheet]

        # Sanity: the ITEM# cell for this block (col A on left, col F on right)
        # must equal the item# we think we're writing for. Catches stale row
        # indices from a catalog that's drifted out of sync.
        item_col = "A" if w.side == "left" else "F"
        actual_item = ws[f"{item_col}{w.row_index}"].value
        if _to_int(actual_item) != w.item_no:
            raise ValueError(
                f"row drift in sheet {w.sheet!r}: expected item#{w.item_no} at "
                f"{item_col}{w.row_index}, found {actual_item!r}"
            )

        cell = ws[f"{ORDER_COL[w.side]}{w.row_index}"]
        cell.value = w.qty
        written += 1

    wb.save(output_path)
    return written


def _to_int(v) -> int | None:
    if v is None or v == "":
        return None
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return None
