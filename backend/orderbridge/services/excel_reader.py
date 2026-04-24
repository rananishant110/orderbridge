"""Excel parsers for OneStop and GrainMarket order forms.

These are READ-ONLY. Any writes go through ``excel_writer`` so we can reason
about formatting preservation in one place.
"""
from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Iterator

from openpyxl import load_workbook

from .normalize import normalize


# ---------- OneStop ---------------------------------------------------------

@dataclass(frozen=True)
class OnestopRow:
    row_index: int
    qty: int
    description: str
    description_normalized: str
    price: float | None
    is_header: bool


def read_onestop(path: Path | str, only_with_qty: bool = False) -> list[OnestopRow]:
    """OneStop has a single sheet ``Report`` laid out as:

        col A = QTY, col B = description, col C = price

    Category headers appear in col B with A + C empty. We surface them so the
    catalog diff can keep section context even if the UI chooses to hide them.
    """
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb["Report"] if "Report" in wb.sheetnames else wb.active

    rows: list[OnestopRow] = []
    for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        # Pad to at least 3 columns
        row = (list(row) + [None, None, None])[:3]
        qty_raw, desc_raw, price_raw = row

        desc = (str(desc_raw).strip() if desc_raw is not None else "")
        if not desc:
            continue

        qty = _coerce_int(qty_raw) or 0
        price = _coerce_float(price_raw)
        is_header = qty_raw in (None, "") and price_raw in (None, "")

        if only_with_qty and qty <= 0:
            continue

        rows.append(
            OnestopRow(
                row_index=idx,
                qty=qty,
                description=desc,
                description_normalized=normalize(desc),
                price=price,
                is_header=is_header,
            )
        )
    return rows


# ---------- GrainMarket ----------------------------------------------------

GM_PRODUCT_SHEETS = (
    "REST LIST",
    "BULK PRODUCTS",
    "GRAIN MARKET PRODUCTS",
    "ORGANIC PRODUCTS",
    "BRANDED PRODUCTS",
    "FROZEN PRODUCTS",
    "NON-FOOD PRODUCTS",
    "SUPPLIES",
)


@dataclass(frozen=True)
class GmRow:
    item_no: int
    sheet: str
    side: str           # "left" or "right"
    row_index: int      # sheet row (1-based)
    description: str
    description_normalized: str
    price: float | None
    order_cell_value: object  # raw existing ORDER value (NA / None / int)
    available: bool


def read_gm_catalog(path: Path | str) -> list[GmRow]:
    """GM uses a two-column side-by-side grid on each product sheet:

        Left side : A=ITEM#, B=DESCRIPTION, C=ORDER, D=PRICE
        Right side: F=ITEM#, G=DESCRIPTION, H=ORDER, I=PRICE
    """
    wb = load_workbook(path, data_only=True, read_only=True)
    rows: list[GmRow] = []
    for sheet_name in GM_PRODUCT_SHEETS:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        rows.extend(_read_gm_sheet(ws, sheet_name))
    return rows


def _read_gm_sheet(ws, sheet_name: str) -> Iterator[GmRow]:
    # read_only mode iterates all cells; iterate by row and pluck by column
    # letter mapping to be robust across sheets with slightly different header rows.
    for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
        # Normalize row length
        row = (list(row) + [None] * 9)[:9]
        # Left block: A,B,C,D = index 0,1,2,3
        left = _build_gm_row(sheet_name, "left", idx,
                             item=row[0], desc=row[1], order=row[2], price=row[3])
        if left is not None:
            yield left
        # Right block: F,G,H,I = index 5,6,7,8
        right = _build_gm_row(sheet_name, "right", idx,
                              item=row[5], desc=row[6], order=row[7], price=row[8])
        if right is not None:
            yield right


def _build_gm_row(sheet, side, row_index, *, item, desc, order, price) -> GmRow | None:
    item_no = _coerce_int(item)
    if item_no is None:
        return None  # header rows, spacers, blanks
    description = (str(desc).strip() if desc is not None else "")
    if not description:
        return None
    order_val = order
    available = not (isinstance(order_val, str) and order_val.strip().upper() == "NA")
    return GmRow(
        item_no=item_no,
        sheet=sheet,
        side=side,
        row_index=row_index,
        description=description,
        description_normalized=normalize(description),
        price=_coerce_float(price),
        order_cell_value=order_val,
        available=available,
    )


# ---------- helpers --------------------------------------------------------

def _coerce_int(value) -> int | None:
    if value is None or value == "":
        return None
    try:
        n = int(float(value))
        return n
    except (TypeError, ValueError):
        return None


def _coerce_float(value) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None
