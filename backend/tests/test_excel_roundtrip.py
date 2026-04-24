"""Verify excel_writer preserves sheet metadata when writing the ORDER column."""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook

from orderbridge.services.excel_reader import read_gm_catalog
from orderbridge.services.excel_writer import OrderWrite, write_quantities


def _build_synthetic_gm(path: Path) -> None:
    """Produce a tiny multi-sheet GM workbook shaped like the real one."""
    wb = Workbook()
    # Remove default, then add expected sheet names
    wb.remove(wb.active)
    for name in ("BRANDED PRODUCTS", "REST LIST"):
        ws = wb.create_sheet(name)
        # header row
        ws.append(["ITEM#", "ITEM DESCRIPTION", "ORDER", "PRICE",
                   None,
                   "ITEM#", "ITEM DESCRIPTION", "ORDER", "PRICE"])
        ws.append([100, "WIDGET 6X1KG", None, 12.50,
                   None,
                   101, "GADGET 12X500G", "NA", 8.25])
        ws.append([102, "THINGAMAJIG 4X2L", None, 5.00,
                   None,
                   None, None, None, None])
    wb.save(path)


def test_writer_sets_order_and_roundtrips(tmp_path: Path) -> None:
    template = tmp_path / "gm_template.xlsx"
    out = tmp_path / "gm_output.xlsx"
    _build_synthetic_gm(template)

    rows = read_gm_catalog(template)
    # Two sheets × (2 left items + 1 right item) = 6
    assert len(rows) == 6

    target = next(r for r in rows if r.item_no == 100 and r.sheet == "BRANDED PRODUCTS")
    n = write_quantities(
        template, out,
        [OrderWrite(sheet=target.sheet, side=target.side,
                    row_index=target.row_index, item_no=100, qty=7)],
    )
    assert n == 1

    wb = load_workbook(out)
    assert wb["BRANDED PRODUCTS"][f"C{target.row_index}"].value == 7
    # Sibling cells untouched
    assert wb["BRANDED PRODUCTS"][f"A{target.row_index}"].value == 100
    assert wb["BRANDED PRODUCTS"][f"B{target.row_index}"].value == "WIDGET 6X1KG"


def test_writer_rejects_stale_row(tmp_path: Path) -> None:
    template = tmp_path / "gm_template.xlsx"
    out = tmp_path / "gm_output.xlsx"
    _build_synthetic_gm(template)

    import pytest
    with pytest.raises(ValueError, match="row drift"):
        write_quantities(
            template, out,
            [OrderWrite(sheet="BRANDED PRODUCTS", side="left",
                        row_index=2, item_no=99999, qty=1)],
        )
